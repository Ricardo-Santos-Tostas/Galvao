#!/usr/bin/env python3
"""
Converte planilha Excel exportada do Access para CSV compatível com o MySQL.

Uso:
  python scripts/converter_xlsx.py "caminho/planilha.xlsx"
  python scripts/converter_xlsx.py "caminho/planilha.xlsx" import/planilha_access.csv
"""

from __future__ import annotations

import csv
import sys
from datetime import date, datetime, time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

# Ordem das colunas no Excel do Access (Planilha1)
COLUNAS_MYSQL = [
    "CADASTRO",
    "RECLAMANTE",
    "DATA_NASC",
    "ENDERE_O",
    "FONE_RTE",
    "FONE_RTE_2_",
    "FONE_RTE_3_",
    "FONE_RTE_4_",
    "FALAR_COM_FONE_1_",
    "FALAR_COM_FONE_2_",
    "FALAR_COM_FONE_3_",
    "FALAR_COM_FONE_4_",
    "RECLAMADA",
    "END_RDA",
    "JUNTA",
    "PROC",
    "DIA_AUD",
    "HORA_AUD",
    "PRA_A_DIA",
    "PRA_A_HORA",
    "ANDAMENTO",
    "CTPS",
    "IDENTIDADE",
    "CPF",
    "COL_2__RECLAMADA",
    "END_RDA_1",
    None,  # 2ª RECLAMANTE — não existe no MySQL
    None,  # END-RTE — não existe no MySQL
    "cxpra_a",
]


def formatar_celula(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        if valor.hour == 0 and valor.minute == 0 and valor.second == 0:
            return valor.strftime("%d/%m/%Y")
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def converter(xlsx_path: Path, csv_path: Path) -> int:
    if not xlsx_path.exists():
        print(f"ERRO: Arquivo não encontrado: {xlsx_path}")
        return 1

    print(f"Lendo: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    nome_aba = "Planilha1" if "Planilha1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[nome_aba]

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    linhas = 0

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([c for c in COLUNAS_MYSQL if c is not None])

        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                continue
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            saida = []
            for col_idx, col_mysql in enumerate(COLUNAS_MYSQL):
                if col_mysql is None:
                    continue
                valor = row[col_idx] if col_idx < len(row) else None
                saida.append(formatar_celula(valor))

            if not saida[0]:
                continue

            writer.writerow(saida)
            linhas += 1

            if linhas % 2000 == 0:
                print(f"  ... {linhas} linhas")

    wb.close()
    print(f"\nCSV gerado: {csv_path}")
    print(f"Total de registros: {linhas}")
    return 0


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1

    xlsx = Path(sys.argv[1]).expanduser().resolve()
    base = Path(__file__).resolve().parent.parent

    if len(sys.argv) >= 3:
        csv_out = Path(sys.argv[2])
        if not csv_out.is_absolute():
            csv_out = base / csv_out
    else:
        csv_out = base / "import" / "planilha_access.csv"

    return converter(xlsx, csv_out)


if __name__ == "__main__":
    raise SystemExit(main())
